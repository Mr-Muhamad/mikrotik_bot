"""Tests for core/reports_excel.py — Excel report generation."""

import io

import openpyxl


def _usage_rows(count=1):
    return [
        {
            "name": f"user{i}",
            "profile": "Basic",
            "status": "active",
            "bytes_in": i * 1000,
            "bytes_out": i * 500,
            "total_bytes": i * 1500,
            "limit_str": "1GB",
            "percent": float(i),
            "comment": f"note{i}",
        }
        for i in range(count)
    ]


def _sales_rows(count=1):
    return [
        {
            "id": str(i),
            "name": f"Batch{i}",
            "profile": "Basic",
            "batch_type": "random",
            "count": i * 10,
            "created_at": f"2025-01-0{i + 1}",
        }
        for i in range(count)
    ]


class TestBuildUsageExcelReport:
    def test_returns_valid_xlsx_bytes(self):
        from core.reports_excel import build_usage_excel_report

        result = build_usage_excel_report({"rows": _usage_rows(1)})
        assert isinstance(result, bytes)
        assert len(result) > 100
        wb = openpyxl.load_workbook(filename=io.BytesIO(result))
        ws = wb.active
        assert ws is not None
        assert ws["A1"].value is not None

    def test_empty_rows_returns_valid_xlsx(self):
        from core.reports_excel import build_usage_excel_report

        result = build_usage_excel_report({"rows": []})
        assert isinstance(result, bytes)
        wb = openpyxl.load_workbook(filename=io.BytesIO(result))
        assert wb.active is not None

    def test_no_rows_key_returns_valid_xlsx(self):
        from core.reports_excel import build_usage_excel_report

        result = build_usage_excel_report({})
        assert isinstance(result, bytes)

    def test_custom_title(self):
        from core.reports_excel import build_usage_excel_report

        result = build_usage_excel_report(
            {"rows": []}, title="تقرير خاص"
        )
        assert isinstance(result, bytes)
        wb = openpyxl.load_workbook(filename=io.BytesIO(result))
        assert "تقرير خاص" in str(wb.active["A1"].value)

    def test_multiple_rows_zebra_striping(self):
        from core.reports_excel import build_usage_excel_report

        result = build_usage_excel_report(
            {"rows": _usage_rows(5)}
        )
        assert isinstance(result, bytes)


class TestBuildSalesExcelReport:
    def test_returns_valid_xlsx_bytes(self):
        from core.reports_excel import build_sales_excel_report

        result = build_sales_excel_report(_sales_rows(1))
        assert isinstance(result, bytes)
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(result)
        )
        assert wb.active is not None

    def test_empty_batches_returns_valid_xlsx(self):
        from core.reports_excel import build_sales_excel_report

        result = build_sales_excel_report([])
        assert isinstance(result, bytes)

    def test_custom_title(self):
        from core.reports_excel import build_sales_excel_report

        result = build_sales_excel_report(
            [], title="مبيعات مخصصة"
        )
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(result)
        )
        title_val = str(wb.active["A1"].value)
        assert "مبيعات مخصصة" in title_val

    def test_multiple_batches(self):
        from core.reports_excel import build_sales_excel_report

        result = build_sales_excel_report(_sales_rows(4))
        assert isinstance(result, bytes)


class TestAutoFitColumns:
    def test_columns_auto_adjusted(self):
        from core.reports_excel import _auto_fit_columns

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["short"])
        ws.append(["a very long column header text"])
        _auto_fit_columns(ws)
        assert ws.column_dimensions["A"].width >= 12
