"""Excel (.xlsx) report generator module using openpyxl.

Generates formatted Excel workbooks with custom header styling,
zebra row fills, and auto-adjusted column widths for Telegram downloads.
"""

from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Style Constants
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")

TITLE_FILL = PatternFill(start_color="0F243E", end_color="0F243E", fill_type="solid")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")

ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def _auto_fit_columns(ws: Any) -> None:
    """Adjust column widths dynamically based on max cell text length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def build_usage_excel_report(
    report: dict[str, Any], title: str = "تقرير استخدام الشبكة"
) -> bytes:
    """Generate a formatted Excel (.xlsx) workbook bytes for hotspot usage report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "تقرير الاستخدام"
    ws.views.sheetView[0].rightToLeft = True

    # Title Banner
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 {title}"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    # Headers
    headers = [
        "اسم المستخدم",
        "الباقة",
        "الحالة",
        "المُحمل (Bytes In)",
        "المُرسل (Bytes Out)",
        "الإجمالي",
        "الحد المسموح",
        "النسبة %",
        "الملاحظات",
    ]
    ws.append([])  # Spacer row 2
    ws.append(headers)  # Row 3
    ws.row_dimensions[3].height = 25

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Data Rows
    rows_data = report.get("rows", [])
    for idx, r in enumerate(rows_data, start=4):
        row_values = [
            r.get("name", ""),
            r.get("profile", ""),
            r.get("status", ""),
            r.get("bytes_in", 0),
            r.get("bytes_out", 0),
            r.get("total_bytes", 0),
            r.get("limit_str", "—"),
            f"{r.get('percent', 0.0):.1f}%",
            r.get("comment", ""),
        ]
        ws.append(row_values)
        ws.row_dimensions[idx].height = 20
        fill = ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL

        for c_idx in range(1, len(row_values) + 1):
            c = ws.cell(row=idx, column=c_idx)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = RIGHT_ALIGN if c_idx in (4, 5, 6) else CENTER_ALIGN

    _auto_fit_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_sales_excel_report(
    batches: list[dict[str, Any]], title: str = "تقرير مبيعات الكروت"
) -> bytes:
    """Generate a formatted Excel (.xlsx) workbook bytes for card batch sales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "المبيعات"
    ws.views.sheetView[0].rightToLeft = True

    # Title Banner
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"💳 {title}"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    # Headers
    headers = [
        "المعرف (ID)",
        "اسم الدفعة",
        "الباقة (Profile)",
        "النوع",
        "عدد الكروت",
        "تاريخ الإنشاء",
    ]
    ws.append([])  # Spacer row 2
    ws.append(headers)  # Row 3
    ws.row_dimensions[3].height = 25

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Data Rows
    for idx, b in enumerate(batches, start=4):
        row_values = [
            b.get("id", ""),
            b.get("name", ""),
            b.get("profile", ""),
            b.get("batch_type", ""),
            b.get("count", 0),
            b.get("created_at", ""),
        ]
        ws.append(row_values)
        ws.row_dimensions[idx].height = 20
        fill = ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL

        for c_idx in range(1, len(row_values) + 1):
            c = ws.cell(row=idx, column=c_idx)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = CENTER_ALIGN

    _auto_fit_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
